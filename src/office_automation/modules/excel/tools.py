"""Excel 模块 — 核心工具函数"""
from pathlib import Path
from office_automation.shared.utils import output_path, timestamp


# --------------- 内部辅助 ---------------

def _create_workbook():
    """创建工作簿"""
    from openpyxl import Workbook
    return Workbook()


def _apply_header_style(ws, headers, style: str, header_row: int = 1):
    """给表头行应用样式

    professional: 深蓝底白字，加粗，居中对齐
    simple: 灰底黑字，加粗
    colorful: 彩色渐变底，白字，加粗

    header_row: 表头所在行号（默认第1行，报表生成时传第2行因为有标题行）
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if style == "professional":
        fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    elif style == "colorful":
        fill = PatternFill(start_color="E74C3C", end_color="C0392B", fill_type="solid")
        font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    else:  # simple
        fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        font = Font(name="微软雅黑", bold=True, color="1F3864", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if header is not None:
            cell.value = header
        cell.font = font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border


def _add_chart_to_sheet(ws, chart_type: str, data_start_row: int, data_end_row: int, num_cols: int):
    """向工作表添加图表"""
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    chart = {
        "bar": BarChart(),
        "line": LineChart(),
        "pie": PieChart(),
    }.get(chart_type, BarChart())

    chart.title = "数据图表"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    if chart_type == "pie":
        cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
    else:
        for col_idx in range(2, num_cols + 1):
            data = Reference(ws, min_col=col_idx, min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        chart.set_categories(cats)

    chart_col = num_cols + 2
    ws.add_chart(chart, f"{_col_letter(chart_col)}2")


def _col_letter(n: int) -> str:
    """数字转 Excel 列字母（1→A, 2→B, ...）"""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


# --------------- 公开工具函数 ---------------

def create_workbook() -> dict:
    """创建一个空白工作簿并保存到输出目录"""
    try:
        wb = _create_workbook()
        wb.save(output_path("new_workbook.xlsx"))
        return {"success": True, "message": "空白工作簿已创建", "files": [str(output_path("new_workbook.xlsx"))]}
    except Exception as e:
        return {"success": False, "error": str(e)}


def add_sheet(filepath: str, sheet_name: str) -> dict:
    """向已有工作簿添加新工作表"""
    try:
        from openpyxl import load_workbook
        p = Path(filepath)
        if not p.exists():
            return {"success": False, "error": f"文件不存在: {filepath}"}
        wb = load_workbook(filepath)
        wb.create_sheet(title=sheet_name)
        wb.save(filepath)
        return {"success": True, "message": f"工作表 '{sheet_name}' 已添加", "files": [filepath]}
    except Exception as e:
        return {"success": False, "error": str(e)}


def format_header(filepath: str, sheet_name: str, style: str = "professional") -> dict:
    """设置已存在工作表的表头样式"""
    try:
        from openpyxl import load_workbook
        p = Path(filepath)
        if not p.exists():
            return {"success": False, "error": f"文件不存在: {filepath}"}
        wb = load_workbook(filepath)
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        _apply_header_style(ws, headers, style)
        wb.save(filepath)
        return {"success": True, "message": f"表头已应用 '{style}' 样式", "files": [filepath]}
    except Exception as e:
        return {"success": False, "error": str(e)}


def add_data_rows(filepath: str, sheet_name: str, data: list[list], start_row: int = 2) -> dict:
    """向工作表添加数据行"""
    try:
        from openpyxl import load_workbook
        p = Path(filepath)
        if not p.exists():
            return {"success": False, "error": f"文件不存在: {filepath}"}
        wb = load_workbook(filepath)
        ws = wb[sheet_name]
        for r_idx, row in enumerate(data, start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(filepath)
        return {"success": True, "message": f"已添加 {len(data)} 行数据", "files": [filepath]}
    except Exception as e:
        return {"success": False, "error": str(e)}


def add_chart(filepath: str, sheet_name: str, chart_type: str = "bar") -> dict:
    """为工作表数据区域添加图表"""
    try:
        from openpyxl import load_workbook
        p = Path(filepath)
        if not p.exists():
            return {"success": False, "error": f"文件不存在: {filepath}"}
        wb = load_workbook(filepath)
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 3:
            return {"success": False, "error": "数据行不足（至少需要表头+2行数据）"}
        _add_chart_to_sheet(ws, chart_type, 2, max_row, max_col)
        wb.save(filepath)
        return {"success": True, "message": f"已添加 {chart_type} 图表", "files": [filepath]}
    except Exception as e:
        return {"success": False, "error": str(e)}


def save(wb_or_filepath, filepath: str = "") -> dict:
    """保存工作簿"""
    try:
        from openpyxl import Workbook, load_workbook
        out = output_path(filepath) if filepath else output_path(f"saved_{timestamp()}.xlsx")
        if isinstance(wb_or_filepath, Workbook):
            wb_or_filepath.save(out)
        elif isinstance(wb_or_filepath, str):
            wb = load_workbook(wb_or_filepath)
            wb.save(out)
        else:
            wb_or_filepath.save(out)
        return {"success": True, "message": "已保存", "files": [str(out)]}
    except Exception as e:
        return {"success": False, "error": str(e)}


def read_excel(filepath: str, sheet_name: str = "") -> dict:
    """读取 Excel 文件内容返回为列表"""
    try:
        from openpyxl import load_workbook
        p = Path(filepath)
        if not p.exists():
            return {"success": False, "error": f"文件不存在: {filepath}"}
        wb = load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        return {"success": True, "message": f"已读取 {len(data)} 行", "data": data}
    except Exception as e:
        return {"success": False, "error": str(e)}


def clean_data(
    source_path: str = "",
    sheet_name: str = "Sheet1",
    remove_duplicates: bool = True,
    fill_empty: str = "",
    trim_whitespace: bool = True,
    drop_empty_rows: bool = True,
    output_filename: str = "cleaned_data.xlsx",
) -> dict:
    """清洗 Excel 数据：去重、填白、修正格式、删空行"""
    try:
        from openpyxl import load_workbook
        p = Path(source_path) if source_path else None
        if not p or not p.exists():
            return {"success": False, "error": f"源文件不存在: {source_path}"}

        wb = load_workbook(source_path)
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        # 读取所有数据
        all_data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            all_data.append(list(row))

        if not all_data:
            return {"success": False, "error": "工作表为空"}

        headers = all_data[0]
        rows = all_data[1:]

        # 处理每行
        cleaned_rows = []
        seen = set()
        for row in rows:
            # trim 空白
            processed = []
            for val in row:
                if isinstance(val, str):
                    processed.append(val.strip() if trim_whitespace else val)
                else:
                    processed.append(val)

            # 填白
            if fill_empty:
                processed = [v if v is not None and str(v).strip() != "" else fill_empty for v in processed]

            # 去重
            if remove_duplicates:
                key = tuple(processed)
                if key in seen:
                    continue
                seen.add(key)

            # 删全空行
            if drop_empty_rows and all(v is None or str(v).strip() == "" for v in processed):
                continue

            cleaned_rows.append(processed)

        # 写回新工作表
        ws_clean = wb.create_sheet(title=f"{sheet_name}_cleaned")
        for c_idx, h in enumerate(headers, 1):
            ws_clean.cell(row=1, column=c_idx, value=h)
        for r_idx, row in enumerate(cleaned_rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws_clean.cell(row=r_idx, column=c_idx, value=val)

        out_path = str(output_path(output_filename))
        wb.save(out_path)
        return {
            "success": True,
            "message": f"数据清洗完成，原始 {len(rows)} 行 → 清洗后 {len(cleaned_rows)} 行",
            "files": [out_path],
            "stats": {"original_rows": len(rows), "cleaned_rows": len(cleaned_rows)},
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def merge_sheets(
    source_path: str = "",
    sheet_names: list[str] = None,
    include_header: bool = True,
    add_source_column: bool = False,
    output_filename: str = "merged_sheets.xlsx",
) -> dict:
    """合并多个工作表为一个"""
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        p = Path(source_path)
        if not p.exists():
            return {"success": False, "error": f"源文件不存在: {source_path}"}

        wb = load_workbook(source_path, data_only=True)
        sheets = sheet_names if sheet_names else wb.sheetnames

        if not sheets:
            return {"success": False, "error": "没有可合并的工作表"}

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "合并结果"
        current_row = 1

        header = None
        for s_name in sheets:
            ws = wb[s_name]
            data = list(ws.iter_rows(values_only=True))
            if not data:
                continue

            if header is None:
                header = list(data[0])
                if add_source_column:
                    header = ["来源工作表"] + header
                for c_idx, h in enumerate(header, 1):
                    cell = out_ws.cell(row=current_row, column=c_idx, value=h)
                    cell.font = Font(bold=True)
                current_row += 1
            elif not include_header:
                data = data[1:]  # 跳过表头

            for row in data[1:]:
                row_data = list(row)
                if add_source_column:
                    row_data = [s_name] + row_data
                for c_idx, val in enumerate(row_data, 1):
                    out_ws.cell(row=current_row, column=c_idx, value=val)
                current_row += 1

        out_path = str(output_path(output_filename))
        out_wb.save(out_path)
        return {
            "success": True,
            "message": f"已合并 {len(sheets)} 个工作表，共 {current_row - 1} 行",
            "files": [out_path],
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def analyze(
    title: str = "数据分析",
    sheet_name: str = "分析结果",
    headers: list = None,
    data: list = None,
    include_chart: bool = True,
    chart_type: str = "bar",
    style: str = "professional",
) -> dict:
    """生成分析报表（generate_report 的别名，带默认图表）"""
    return generate_report(
        title=title,
        sheet_name=sheet_name,
        headers=headers or [],
        data=data or [],
        include_chart=include_chart,
        chart_type=chart_type,
        style=style,
    )


def generate_report(
    title: str = "报表",
    sheet_name: str = "Sheet1",
    headers: list = None,
    data: list = None,
    include_chart: bool = False,
    chart_type: str = "bar",
    style: str = "professional",
) -> dict:
    """生成专业报表（完整流程：创建→写表头→填数据→加图表→保存）"""
    try:
        headers = headers or []
        data = data or []
        wb = _create_workbook()
        ws = wb.active
        ws.title = sheet_name

        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        from openpyxl.styles import Font, Alignment
        title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="2F5496")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 表头（第2行）
        for c_idx, h in enumerate(headers, 1):
            ws.cell(row=2, column=c_idx, value=h)
        _apply_header_style(ws, headers, style, header_row=2)

        # 数据（从第3行开始）
        for r_idx, row in enumerate(data, 3):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # 图表
        if include_chart and len(data) >= 1 and len(headers) >= 2:
            data_end_row = 2 + len(data)
            _add_chart_to_sheet(ws, chart_type, 3, data_end_row, len(headers))

        # 保存
        filename = f"report_{timestamp()}.xlsx"
        out = str(output_path(filename))
        wb.save(out)
        return {
            "success": True,
            "message": f"报表已生成: {title}",
            "files": [out],
            "stats": {"rows": len(data), "columns": len(headers), "has_chart": include_chart},
        }
    except Exception as e:
        return {"success": False, "error": str(e)}
